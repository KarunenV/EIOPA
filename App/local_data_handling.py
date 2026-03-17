import os
import pandas as pd
from openpyxl import load_workbook, Workbook


def edit_local_Excel(results: dict, config: dict):
    TARGET_SHEETS = config.get("target_sheets", [])
    FINANCES = config.get("finances", [])
    OUTPUT_DIR = config.get("output_dir", "output_data")

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    for finance in FINANCES:
        for sheet in TARGET_SHEETS:

            out_name = os.path.join(OUTPUT_DIR, f"{finance}.xlsx")
            print(f"Saving {finance} data to {out_name} (sheet: {sheet})")
            results_key = f"{finance}_{sheet}"
            if results_key not in results:
                print(f"   ⚠️ No data found for {results_key}, skipping.")
                continue

            results_df = results[results_key]

            # date header to match is first column value of the first row in results_df
            date_to_match = str(results_df.iloc[0, 0]).strip() if not results_df.empty else None
            # data to write starts after date header row
            results_data = results_df.iloc[1:].reset_index(drop=True)

            print(f"   → Writing {len(results_data)} rows and {len(results_data.columns)} columns to {out_name} (date {date_to_match})")

            target_col = 2
            if os.path.exists(out_name):
                wb = load_workbook(out_name)
                if sheet in wb.sheetnames:
                    ws = wb[sheet]
                    top_row_values = [str(cell.value).strip() for cell in ws[1] if cell.value is not None and str(cell.value).strip() != ""]

                    if date_to_match and date_to_match in top_row_values:
                        print(f"   → Date {date_to_match} already exists in top row for {sheet}; skipping data write")
                        wb.close()
                        continue

                    # find first empty top cell from column 2 to the right
                    for idx, cell in enumerate(ws[1], start=1):
                        if idx < 2:
                            continue
                        if cell.value is None or str(cell.value).strip() == "":
                            target_col = idx
                            break
                    else:
                        target_col = max(2, len(ws[1]) + 1)

                    if date_to_match:
                        ws.cell(row=1, column=target_col, value=date_to_match)
                    wb.save(out_name)
                    wb.close()
                else:
                    ws = wb.create_sheet(sheet)
                    if date_to_match:
                        ws.cell(row=1, column=2, value=date_to_match)
                    wb.save(out_name)
                    wb.close()
                    target_col = 2
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = sheet
                if date_to_match:
                    ws.cell(row=1, column=2, value=date_to_match)
                wb.save(out_name)
                wb.close()
                target_col = 2

            # write data block under date column
            if not results_data.empty:
                with pd.ExcelWriter(out_name, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                    results_data.to_excel(
                        writer,
                        sheet_name=sheet,
                        index=False,
                        header=False,
                        startrow=1,
                        startcol=target_col - 1,
                    )
            else:
                print("   → No data rows to write after date header")


